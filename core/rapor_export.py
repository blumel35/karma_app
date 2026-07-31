"""
core/rapor_export.py

Talep Merkezi, Portföy Merkezi ve Arşiv Merkezi'nde ortak kullanılan
Excel export modülü. Eski Tkinter uygulamasındaki (talep_ekrani.py)
"filtreli_listeyi_excel_aktar" mantığının Streamlit'e taşınmış hali:

- Aynı görsel kimlik (navy/gold başlık, biçimlendirilmiş tablo)
- "Rapor" (liste) + "Mail_Detay" sayfaları
- FARK: Tkinter yerel diske (Downloads) yazıyordu; Streamlit sunucu
  tarafında BytesIO üretir, kullanıcı st.download_button ile kendi
  bilgisayarına indirir.
- YENİ: İşlem Tipi'ne göre (Satılık / Kiralık / Belirsiz) otomatik
  ayrı sayfalara bölünür — tek dosyada iki liste.

Kullanım (bir sayfadan):

    from core.rapor_export import excel_raporu_olustur, export_butonu_goster

    export_butonu_goster(
        kayitlar=f,                      # o anki filtrelenmiş liste (favori ilçe HARİÇ)
        rapor_basligi="Talep Raporu",
        kayit_tipi="talep",              # "talep" | "portfoy"
        dosya_on_eki="talep_raporu",
    )
"""

from io import BytesIO
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1B2A4A"
GOLD = "C9A84C"
WHITE = "FFFFFF"
MID_GRAY = "DDE1EB"
DARK_GRAY = "4A5568"
SOFT_BLUE = "EAF0FB"
CARD_BG = "F0F4FF"


def _islem_tipi_norm(v):
    ham = str(v.get("islem_tipi") or "").strip()
    if not ham:
        return "Belirsiz"
    low = ham.lower().replace("ı", "i").replace("İ", "i")
    if "kirali" in low:
        return "Kiralık"
    if "satili" in low:
        return "Satılık"
    return "Belirsiz"


def _isim_ayikla(g):
    """'Ad Soyad <mail@adres.com>' -> 'Ad Soyad'"""
    if not g:
        return ""
    g = str(g)
    if "<" in g:
        return g.split("<")[0].strip().strip('"')
    return g.strip()


def _tarih_kisalt(s):
    if not s:
        return ""
    try:
        from email.utils import parsedate_to_datetime
        d = parsedate_to_datetime(s)
        return d.strftime("%d.%m.%Y")
    except Exception:
        return str(s)[:10]


def _talep_satiri(v):
    return {
        "Kayıt Tarihi": _tarih_kisalt(v.get("kayit_tarihi", "")),
        "İlçe": ", ".join(v.get("ilceler") or []) or (v.get("ilce") or "") or "-",
        "Bölge / Mahalle": v.get("bolge_mahalle") or "-",
        "Mülk Tipi": v.get("mulk_tipi") or "Belirsiz",
        "Oda / m²": v.get("oda_sayisi_m2") or "-",
        "Bütçe": v.get("max_butce") or "-",
        "Özet": (v.get("ozet") or v.get("ozel_kriterler") or "")[:200],
        "Danışman": _isim_ayikla(v.get("talep_eden_danisan", "")) or "-",
    }


def _portfoy_satiri(v):
    return {
        "Kayıt Tarihi": _tarih_kisalt(v.get("kayit_tarihi", "")),
        "İlçe": ", ".join(v.get("ilceler") or []) or (v.get("ilce") or "") or "-",
        "Bölge / Mahalle": v.get("bolge_mahalle") or "-",
        "Mülk Tipi": v.get("mulk_tipi") or "Belirsiz",
        "Oda / m²": v.get("oda_sayisi_m2") or "-",
        "Fiyat": v.get("fiyat") or "-",
        "Özet": (v.get("ozet") or v.get("ozellikler") or "")[:200],
        "Danışman": _isim_ayikla(v.get("talep_eden_danisan", "")) or "-",
    }


import re


def _html_temizle(text):
    """
    Veritabanındaki mail_icerigi alanı bazı kayıtlarda (özellikle zincir/
    alıntılı yazışmalarda) hiç temizlenmemiş HTML içerebiliyor — mail_fetcher
    tarafındaki strip_html_tags her durumda çalışmamış olabilir. Export'a
    ham HTML'in sızmaması için burada savunma amaçlı ikinci bir temizlik
    katmanı uygulanıyor.
    """
    if not text:
        return ""
    text = str(text)
    text = re.sub(r"<style.*?>.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    text = re.sub(r"<script.*?>.*?</script>", " ", text, flags=re.DOTALL | re.IGNORECASE)
    # <br> ve <div> kapanışları satır sonu anlamına gelir — okunabilirlik
    # için gerçek satır sonuna çeviriyoruz, düz boşluğa değil.
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</div>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    # HTML entity'lerini çöz (&nbsp; &amp; &lt; &gt; &#39; vb.)
    import html as _html
    text = _html.unescape(text)
    # Art arda gelen boş satırları teke indir
    text = re.sub(r"\n\s*\n+", "\n\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def _mail_detay_satiri(v):
    return {
        "Mail Konusu": _html_temizle(v.get("mail_konusu", ""))[:300],
        "Mail İçeriği": _html_temizle(v.get("mail_icerigi", ""))[:3000],
    }


def _sayfa_bicimlendir(ws, baslik, satir_sayisi, sutun_sayisi):
    """Rapor sayfasına navy/gold başlık ve tablo biçimlendirmesi uygular."""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 8

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=sutun_sayisi)
    baslik_hucre = ws.cell(row=2, column=1, value=baslik)
    baslik_hucre.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    baslik_hucre.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    baslik_hucre.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[2].height = 34
    for col in range(1, sutun_sayisi + 1):
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=NAVY)

    tarih_str = datetime.now().strftime("%d.%m.%Y %H:%M")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=sutun_sayisi)
    alt_baslik = ws.cell(row=3, column=1, value=f"Oluşturulma: {tarih_str}  ·  Toplam kayıt: {satir_sayisi}")
    alt_baslik.font = Font(name="Calibri", size=10, italic=True, color=DARK_GRAY)
    alt_baslik.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    alt_baslik.fill = PatternFill("solid", fgColor=SOFT_BLUE)
    ws.row_dimensions[3].height = 20

    baslik_satiri = 5
    ince_kenar = Side(style="thin", color=MID_GRAY)
    for col in range(1, sutun_sayisi + 1):
        hucre = ws.cell(row=baslik_satiri, column=col)
        hucre.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        hucre.fill = PatternFill("solid", fgColor=GOLD)
        hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hucre.border = Border(bottom=ince_kenar)
    ws.row_dimensions[baslik_satiri].height = 26

    for row in range(baslik_satiri + 1, baslik_satiri + 1 + satir_sayisi):
        zebra = CARD_BG if (row - baslik_satiri) % 2 == 0 else WHITE
        for col in range(1, sutun_sayisi + 1):
            hucre = ws.cell(row=row, column=col)
            hucre.fill = PatternFill("solid", fgColor=zebra)
            hucre.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            hucre.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            hucre.border = Border(bottom=ince_kenar)

    for col in range(1, sutun_sayisi + 1):
        harf = get_column_letter(col)
        baslik_metni = str(ws.cell(row=baslik_satiri, column=col).value or "")
        genislik = max(14, min(45, len(baslik_metni) + 8))
        ws.column_dimensions[harf].width = genislik

    ws.freeze_panes = ws.cell(row=baslik_satiri + 1, column=1)


def _detay_sayfa_bicimlendir(ws, rapor_sayfa_adi, satir_sayisi):
    """Mail_Detay sayfasına başlık, geri dön linki ve okunabilir sütun
    genişliği/satır yüksekliği uygular."""
    ws.sheet_view.showGridLines = False

    # Önce boş satırı ekle (header 2. satıra kayar, veri 3.'ten başlar),
    # SONRA geri-dön linkini bu garanti-boş satıra yaz. Tersi sırada
    # (değer/hyperlink önce, insert_rows sonra) openpyxl'de hyperlink'in
    # satırla birlikte kaymaması riski var.
    ws.insert_rows(1)
    baslik_satiri = 2

    geri_hucre = ws.cell(row=1, column=1)
    geri_hucre.value = "← Rapora Dön"
    geri_hucre.hyperlink = f"#'{rapor_sayfa_adi}'!A1"
    geri_hucre.font = Font(name="Calibri", size=10, color="1155CC", underline="single")

    for col in range(1, ws.max_column + 1):
        harf = get_column_letter(col)
        ws.column_dimensions[harf].width = 70 if col == ws.max_column else (10 if col == 1 else 40)

    ince_kenar = Side(style="thin", color=MID_GRAY)
    for col in range(1, ws.max_column + 1):
        hucre = ws.cell(row=baslik_satiri, column=col)
        hucre.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        hucre.fill = PatternFill("solid", fgColor=GOLD)
        hucre.alignment = Alignment(horizontal="center", vertical="center")
        hucre.border = Border(bottom=ince_kenar)

    for row in range(baslik_satiri + 1, baslik_satiri + 1 + satir_sayisi):
        zebra = CARD_BG if (row - baslik_satiri) % 2 == 0 else WHITE
        for col in range(1, ws.max_column + 1):
            hucre = ws.cell(row=row, column=col)
            hucre.fill = PatternFill("solid", fgColor=zebra)
            hucre.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            hucre.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            hucre.border = Border(bottom=ince_kenar)
        ws.row_dimensions[row].height = 90

    ws.freeze_panes = ws.cell(row=baslik_satiri + 1, column=1)


def excel_raporu_olustur(kayitlar, rapor_basligi, kayit_tipi="talep"):
    """
    kayitlar: dict listesi (Supabase satırları)
    kayit_tipi: "talep" | "portfoy" — hangi satır şablonunun kullanılacağını belirler
    Döner: BytesIO (xlsx içeriği)
    """
    satir_fn = _talep_satiri if kayit_tipi == "talep" else _portfoy_satiri

    gruplar = {"Satılık": [], "Kiralık": [], "Belirsiz": []}
    for v in kayitlar:
        gruplar[_islem_tipi_norm(v)].append(v)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        bos_sayfa_yazildi = False
        for grup_adi, grup_kayitlari in gruplar.items():
            if not grup_kayitlari:
                continue
            liste_df = pd.DataFrame([satir_fn(v) for v in grup_kayitlari])
            liste_df.insert(0, "No", range(1, len(liste_df) + 1))
            liste_df["Detay"] = "✉ Detay Aç"
            sayfa_adi = grup_adi[:31]
            liste_df.to_excel(writer, sheet_name=sayfa_adi, index=False, startrow=4)

            detay_df = pd.DataFrame([_mail_detay_satiri(v) for v in grup_kayitlari])
            detay_df.insert(0, "No", range(1, len(detay_df) + 1))
            detay_sayfa_adi = f"{grup_adi[:20]}_Detay"
            detay_df.to_excel(writer, sheet_name=detay_sayfa_adi, index=False)
            bos_sayfa_yazildi = True

        if not bos_sayfa_yazildi:
            pd.DataFrame({"Bilgi": ["Bu filtrede kayıt bulunamadı."]}).to_excel(
                writer, sheet_name="Rapor", index=False
            )

    buffer.seek(0)
    wb = load_workbook(buffer)
    for grup_adi, grup_kayitlari in gruplar.items():
        sayfa_adi = grup_adi[:31]
        detay_sayfa_adi = f"{grup_adi[:20]}_Detay"
        if sayfa_adi in wb.sheetnames and grup_kayitlari:
            ws = wb[sayfa_adi]
            _sayfa_bicimlendir(ws, f"{rapor_basligi} — {grup_adi}", len(grup_kayitlari), ws.max_column)

        if detay_sayfa_adi in wb.sheetnames and grup_kayitlari:
            ws_detay = wb[detay_sayfa_adi]
            _detay_sayfa_bicimlendir(ws_detay, sayfa_adi, len(grup_kayitlari))

            if sayfa_adi in wb.sheetnames:
                # Rapor sayfasındaki her satırın "Detay" hücresine, ilgili
                # kaydın Mail_Detay sayfasındaki satırına giden tıklanabilir
                # bağlantı ekle. Rapor verisi 6. satırdan, Detay verisi
                # (geri-dön satırı eklendiği için) 3. satırdan başlıyor.
                ws_rapor = wb[sayfa_adi]
                detay_sutun = ws_rapor.max_column
                for i in range(len(grup_kayitlari)):
                    rapor_satir = 6 + i
                    detay_satir = 3 + i
                    hucre = ws_rapor.cell(row=rapor_satir, column=detay_sutun)
                    hucre.hyperlink = f"#'{detay_sayfa_adi}'!A{detay_satir}"
                    hucre.font = Font(name="Calibri", size=10, color="1155CC", underline="single")

    cikti = BytesIO()
    wb.save(cikti)
    cikti.seek(0)
    return cikti


def export_butonu_goster(kayitlar, rapor_basligi, kayit_tipi="talep", dosya_on_eki="rapor", key_prefix="exp"):
    """
    Sayfaya 'Excel'e Aktar' butonu ekler. Basılınca dosya üretilir ve
    st.download_button ile indirme linki gösterilir.

    kayitlar: o anda ekrandaki filtreye göre süzülmüş liste (favori ilçe
              kapsamı HARİÇ tutulmalı — bu kişisel bir filtre, rapor
              tüm ilgili ilçeleri içermeli).
    """
    toplam = len(kayitlar)
    col1, col2 = st.columns([1, 3])
    with col1:
        uret = st.button(
            f"📊 Excel'e Aktar ({toplam} kayıt)",
            key=f"{key_prefix}_excel_btn",
            use_container_width=True,
        )
    if uret:
        if toplam == 0:
            st.warning("Aktarılacak kayıt yok — filtreyi genişletmeyi deneyin.")
        else:
            with st.spinner("Excel raporu hazırlanıyor..."):
                dosya = excel_raporu_olustur(kayitlar, rapor_basligi, kayit_tipi)
            dosya_adi = f"{dosya_on_eki}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            with col2:
                st.download_button(
                    label=f"⬇️ {dosya_adi} — indir",
                    data=dosya,
                    file_name=dosya_adi,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{key_prefix}_excel_download",
                    use_container_width=True,
                )
